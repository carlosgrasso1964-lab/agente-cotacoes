# -*- coding: utf-8 -*-
"""Gera graficos de evolucao das cotacoes a partir da planilha.

Le:       cotacoes_historico.xlsx (colunas: Data, Dolar, Euro, Libra, Bitcoin, Ibovespa, Ouro)
Gera:     graficos/<nome>_evolucao.png  — um por variável, em linha.
Uso:      python gerar_graficos.py
"""
import os
from datetime import datetime
import matplotlib
matplotlib.use("Agg")  # sem janela, salva direto em PNG
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook

PLANILHA = "cotacoes_historico.xlsx"
PASTA_OUT = "graficos"

# nome (coluna) -> (título, cor)
VARIAVEIS = [
    ("Dolar",    "Dólar (R$ / US$)",   "#1f77b4"),
    ("Euro",     "Euro (R$ / EUR)",    "#2ca02c"),
    ("Libra",    "Libra (R$ / GBP)",   "#9467bd"),
    ("Bitcoin",  "Bitcoin (em R$)",    "#ff7f0e"),
    ("Ibovespa", "Ibovespa (pontos)",  "#d62728"),
    ("Ouro",     "Ouro (R$/oz)",       "#8c564b"),
]


def ler_dados():
    wb = load_workbook(PLANILHA)
    ws = wb.active
    linhas = [r for r in ws.iter_rows(values_only=True) if r[0]]
    if len(linhas) < 2:
        raise SystemExit("Sem dados suficientes na planilha (só cabeçalho?).")
    header = linhas[0]
    idx = {h: i for i, h in enumerate(header)}
    datas = []
    for r in linhas[1:]:
        # aceita str dd/mm/aaaa ou datetime
        d = r[idx["Data"]]
        if isinstance(d, datetime):
            datas.append(d.date())
        else:
            datas.append(datetime.strptime(str(d), "%d/%m/%Y").date())
    dados = {}
    for nome, _, _ in VARIAVEIS:
        col = idx.get(nome)
        if col is not None:
            dados[nome] = [r[col] for r in linhas[1:]]
    return datas, dados


def gerar():
    datas, dados = ler_dados()
    os.makedirs(PASTA_OUT, exist_ok=True)
    arquivos = []

    for nome, titulo, cor in VARIAVEIS:
        vals = dados.get(nome)
        if not vals or not any(v is not None for v in vals):
            continue
        # remove None
        x = [d for d, v in zip(datas, vals) if v is not None]
        y = [v for v in vals if v is not None]

        fig, ax = plt.subplots(figsize=(9, 4.5), dpi=130)
        ax.plot(x, y, marker="o", markersize=4, linewidth=2, color=cor, label=nome)
        ax.set_title(f"Evolução — {titulo}", fontsize=13, fontweight="bold", pad=12)
        ax.set_xlabel("Data", fontsize=10)
        ax.set_ylabel(titulo, fontsize=10)
        ax.grid(True, linestyle="--", alpha=0.4)
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%d/%m"))
        ax.xaxis.set_major_locator(mdates.AutoDateLocator(minticks=3, maxticks=8))
        fig.autofmt_xdate()
        # formata eixo Y de forma legível
        ax.ticklabel_format(style="plain", axis="y")
        fig.tight_layout()
        arquivo = os.path.join(PASTA_OUT, f"{nome.lower()}_evolucao.png")
        fig.savefig(arquivo, bbox_inches="tight")
        plt.close(fig)
        arquivos.append(arquivo)
        print(f"OK  {arquivo}  ({len(x)} ponto(s))")

    if not arquivos:
        print("Nenhum gráfico gerado.")
    else:
        print(f"\n{len(arquivos)} gráfico(s) gerado(s) em: {os.path.abspath(PASTA_OUT)}")


if __name__ == "__main__":
    gerar()