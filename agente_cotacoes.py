# -*- coding: utf-8 -*-
"""Agente de Cotações — apenas cotações (sem notícias).

- Busca cotações via Yahoo Finance
- Grava na planilha acumulativa (cotacoes_historico.xlsx)
- NÃO busca notícias (isso agora é o journal.py separado)
"""
import requests
import datetime
import os
from openpyxl import Workbook, load_workbook

ARQUIVO_PLANILHA = "cotacoes_historico.xlsx"


def get_cotacoes():
    """Busca cotações atuais e retorna (dados_formatados, numeros_crus)."""
    dados = {
        'ibovespa': 'Indisponivel',
        'dolar': 'Indisponivel',
        'euro': 'Indisponivel',
        'libra': 'Indisponivel',
        'ouro': 'Indisponivel',
        'bitcoin': 'Indisponivel'
    }
    numeros = {}  # valores crus para gravar na planilha/grafico
    h = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'}

    # USD/BRL
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/BRL%3DX?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            usd = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            dados['dolar'] = f"R$ {usd:.2f}"
            dados['usd'] = usd
            numeros['dolar'] = round(float(usd), 4)
    except:
        pass

    # EUR/BRL
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/EURBRL%3DX?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            eur = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            dados['euro'] = f"R$ {eur:.2f}"
            numeros['euro'] = round(float(eur), 4)
    except:
        pass

    # GBP/BRL
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/GBPBRL%3DX?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            gbp = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            dados['libra'] = f"R$ {gbp:.2f}"
            numeros['libra'] = round(float(gbp), 4)
    except:
        pass

    # Bitcoin
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/BTC-USD?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            btc = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            usd = numeros.get('dolar', dados.get('usd', 5.0))
            dados['bitcoin'] = f"R$ {(btc * usd):,.2f}"
            numeros['bitcoin'] = round(float(btc * usd), 2)
    except:
        pass

    # Ibovespa
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/%5EBVSP?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            ibov = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            dados['ibovespa'] = f"{ibov:,.2f} pts"
            numeros['ibovespa'] = round(float(ibov), 2)
    except:
        pass

    # Ouro
    try:
        r = requests.get("https://query1.finance.yahoo.com/v8/finance/chart/GC%3DF?interval=1d&range=1d", headers=h, timeout=15)
        if r.status_code == 200:
            gold = r.json()['chart']['result'][0]['meta']['regularMarketPrice']
            usd = numeros.get('dolar', dados.get('usd', 5.0))
            dados['ouro'] = f"R$ {(gold * usd):,.2f}/oz"
            numeros['ouro'] = round(float(gold * usd), 2)
    except:
        pass

    return dados, numeros


def gravar_historico(numeros):
    """Acrescenta uma linha c/ as cotações na planilha acumulativa (xlsx)."""
    if not numeros:
        print("Nenhuma cotacao numerica disponivel; nada gravado.")
        return

    # Data do ultimo fechamento = dia anterior a execucao (cron roda antes da abertura)
    data_ref = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d/%m/%Y")

    cabecalho = ["Data", "Dolar", "Euro", "Libra", "Bitcoin", "Ibovespa", "Ouro"]
    linha = [
        data_ref,
        numeros.get('dolar', ''),
        numeros.get('euro', ''),
        numeros.get('libra', ''),
        numeros.get('bitcoin', ''),
        numeros.get('ibovespa', ''),
        numeros.get('ouro', ''),
    ]

    try:
        if os.path.exists(ARQUIVO_PLANILHA):
            wb = load_workbook(ARQUIVO_PLANILHA)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cotacoes"
            ws.append(cabecalho)
        ws.append(linha)
        wb.save(ARQUIVO_PLANILHA)
        print(f"Historico gravado em {ARQUIVO_PLANILHA} (linha {ws.max_row})")
    except Exception as e:
        print(f"Erro ao gravar historico: {e}")


if __name__ == "__main__":
    cot, numeros = get_cotacoes()
    gravar_historico(numeros)

    token = os.environ.get("TELEGRAM_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")

    # Modo local (sem Telegram): apenas grava a cotação
    if not token or not chat_id:
        print("Modo local (sem Telegram): cotacao gravada.")